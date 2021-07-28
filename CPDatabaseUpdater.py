def AcquireDependencies():

    pkgNames = {pkg.key for pkg in working_set}
    def InstallPackage(packageNameList):
    #packageNameList:  [spelling for pip (key), spelling for pkg_resources (value)]
        packageAbsent = len({packageNameList[1]} - pkgNames)
        if packageAbsent:
            check_call([executable, "-m", "pip", "install", packageNameList[0]], stdout=DEVNULL)
            print(f"{packageNameList[0]} Installed.")

    InstallPackage(['pikepdf', 'pikepdf'])
    InstallPackage(['PyPDF2', 'pypdf2'])

def AcquirePDFPassword():
    userName = getlogin()
    if '$$$$$$$$$$' in userName:
        return '$$$$$$$$$$'
    elif '$$$$$$$$$$' in userName:
        return '$$$$$$$$$$'
    else:
        name = input('Please Enter Your Full Name (ex. John Doe): ').replace(" ", "").lower()
        with open(DEPENDENCIES_DIR + '$$$$$$$$$$', 'r') as f:
            for line in f.read().split('\n'):
                txtName, password = line.split(' ; ')
                txtName = txtName.strip().replace(" ", "").lower()
                if name == txtName:
                    return password

    print("Name Not Found. Please Enter Your Name And Password Into '$$$$$$$$$$'")
    sleep(3)
    print('Program Exiting Safely')
    sleep(2)
    exit()

def InitializeInbox():
    outlook = Dispatch('outlook.application').GetNamespace('MAPI')
    inbox = outlook.GetDefaultFolder(6)
    inboxFolders = {folder.Name for folder in inbox.Folders}

    if '$$$$$$$$$$' in inboxFolders:
        citiFolder = inbox.Folders.Item('$$$$$$$$$$')
    else:
        inbox.Folders.Add('$$$$$$$$$$')
        citiFolder = inbox.Folders.Item('$$$$$$$$$$')
        print("Please Move All Emails From '$$$$$$$$$$' Into The New '$$$$$$$$$$' Folder in Your Inbox.")
        sleep(4)
        key = input("Once Completed, Please Press Any Key And Then Hit Enter To Continue: ")

def RemoveDateFromCPDatabase(undoDate):
    CPDATABASE_WORKBOOK = load_workbook(ISSUANCE_DB_PATH)
    dateLogSheet = CPDATABASE_WORKBOOK['Date_Log']
    currentRowIndex = dateLogSheet['D1'].value
    undoDatePresent = False
    for i in range(2, currentRowIndex):
        if dateLogSheet['A' + str(i)].value.date() == undoDate:
            dateLogSheet.delete_rows(idx = i, amount = 1)
            currentRowIndex -= 1
            dateLogSheet['D1'].value = currentRowIndex
            undoDatePresent = True
            break

    if undoDatePresent:
        MF_CPReportSheet = CPDATABASE_WORKBOOK['Complete_CP_Report']
        currentRowIndex = MF_CPReportSheet['P1'].value
        i = 2
        while i < currentRowIndex:
            if MF_CPReportSheet['C' + str(i)].value.date() == undoDate:
                MF_CPReportSheet.delete_rows(idx = i, amount = 1)
                currentRowIndex -= 1
            else:
                i += 1
        MF_CPReportSheet['P1'].value = currentRowIndex

        MF_DebtReportSheet = CPDATABASE_WORKBOOK['MF_CP_Report']
        currentRowIndex = MF_DebtReportSheet['M1'].value
        i = 2
        while i < currentRowIndex:
            if MF_DebtReportSheet['B' + str(i)].value.date() == undoDate:
                MF_DebtReportSheet.delete_rows(idx = i, amount = 1)
                currentRowIndex -= 1
            else:
                i += 1
        MF_DebtReportSheet['M1'].value = currentRowIndex

        CPDATABASE_WORKBOOK.save(ISSUANCE_DB_PATH)

    else:
        print(f"There Is No Data For {undoDate}.")
        sleep(2)
        print('Program Exiting Safely')
        sleep(2)
        exit()

def AcquireQueryDate():
    date = None
    while not date:
        response = input("Do You Want To Acquire Data For The Current Date? (y/n/u/e): ")
        if response == 'y':
            date = datetime.today().date()
        elif response == 'n':
            stringDate = input("Please Enter The Day For Which You Wish To Acquire Data (yyyy/mm/dd): ")
            try:
                dateTokens = stringDate.split('/')
                date = datetime(int(dateTokens[0]), int(dateTokens[1]), int(dateTokens[2])).date()
            except:
                print("Please Insert Date In Correct Format.")
                sleep(2)
                date = None
        elif response == 'u':
            stringDate = input("Please Enter The Day For Which You Wish To ERASE Data (yyyy/mm/dd): ")
            try:
                dateTokens = stringDate.split('/')
                undoDate = datetime(int(dateTokens[0]), int(dateTokens[1]), int(dateTokens[2])).date()
            except:
                print("Please Insert Date In Correct Format.")
                sleep(2)
                date = None
            try:
                RemoveDateFromCPDatabase(undoDate)
                print(f"Data For {undoDate} Removed.")
                sleep(2)
                print('Returning To The Main Menu.')
                date = None
            except PermissionError:
                print("The CP Database Is Currently Open. Please Ask Other User(s) To Exit Before Running.")
                sleep(3)
                print('Program Exiting Safely')
                sleep(2)
                exit()
        elif response == 'e':
            print('Program Exiting Safely')
            sleep(2)
            exit()
        else:
            print("Please Answer 'y', 'n', 'u', or 'e'.")
            sleep(2)
            date = None
    return date

def CheckDateLog():

    workBook = load_workbook(ISSUANCE_DB_PATH)
    DateLogSheet = workBook['Date_Log']
    currentRowIndex = DateLogSheet['D1'].value
    for i in range(2, currentRowIndex):
        if DateLogSheet['A' + str(i)].value.date() == CURRENT_DATE:
            print(f"The CP Database Has Already Been Updated For {CURRENT_DATE}.")
            sleep(3)
            print('Program Exiting Safely')
            sleep(2)
            exit()
    return workBook

def AcquireIssuanceReport():

    def AnalyzeIssuanceEmail(message, currentDate):
        analyze = False
        if isinstance(message.Sender.GetExchangeUser(), type(None)):
            if message.SentOn.date() == currentDate:
                if '$$$$$$$$$$' in message.subject:
                    analyze = True
        return analyze

    def AnalyzeDailyBalanceEmail(message, currentDate):
        analyze = False
        if isinstance(message.Sender.GetExchangeUser(), type(None)):
            if message.SentOn.date() == currentDate:
                if '$$$$$$$$$$' in message.subject:
                    analyze = True
        return analyze

    outlook = Dispatch('outlook.application').GetNamespace('MAPI')
    inbox = outlook.GetDefaultFolder(6)
    citiFolder = inbox.Folders.Item('$$$$$$$$$$')
    citiMessages = citiFolder.Items
    citiMessages.sort("[ReceivedTime]", True)
    citiMessage = citiMessages.GetFirst()
    acquiredIssuance = False
    acquiredBalance = False
    for _ in range(50):
        if not isinstance(citiMessage, type(None)):
            if not acquiredIssuance:
                if AnalyzeIssuanceEmail(citiMessage, CURRENT_DATE):
                    for attachment in citiMessage.Attachments:
                        if '$$$$$$$$$$' in attachment.filename:
                            stringDate = '_' + CURRENT_DATE.strftime("%m-%d-%Y")
                            newFileName = attachment.filename[:-4].replace('attachments', 'Issuance') + stringDate + '.pdf'
                            attachment.SaveAsFile(PDF_STORAGE_DIR + attachment.filename)
                            if not isfile(PDF_STORAGE_DIR + newFileName):
                                rename(PDF_STORAGE_DIR + attachment.filename, PDF_STORAGE_DIR + newFileName)
                                print(f"Downloaded File:  {newFileName}")
                            else:
                                remove(PDF_STORAGE_DIR + newFileName)
                                rename(PDF_STORAGE_DIR + attachment.filename, PDF_STORAGE_DIR + newFileName)
                                print(f"Updated File:  {newFileName}")
                            acquiredIssuance = True

            if not acquiredBalance:
                if CURRENT_DATE.weekday() < 5:
                    if AnalyzeDailyBalanceEmail(citiMessage, CURRENT_DATE):
                        for attachment in citiMessage.Attachments:
                            if 'attachments.pdf' in attachment.filename:
                                attachment.SaveAsFile(PDF_STORAGE_DIR + attachment.filename)
                                stringDate = '_' + CURRENT_DATE.strftime("%m-%d-%Y")
                                newFileName = attachment.filename[:-4].replace('attachments', 'DailyBalance') + stringDate + '.pdf'
                                if not isfile(PDF_STORAGE_DIR + newFileName):
                                    rename(PDF_STORAGE_DIR + attachment.filename, PDF_STORAGE_DIR + newFileName)
                                    print(f"Downloaded File:  {newFileName}")
                                else:
                                    remove(PDF_STORAGE_DIR + newFileName)
                                    rename(PDF_STORAGE_DIR + attachment.filename, PDF_STORAGE_DIR + newFileName)
                                    print(f"Updated File:  {newFileName}")
                                acquiredBalance = True

            if acquiredIssuance and acquiredBalance:
                break
            else:
                citiMessage = citiMessages.GetNext()
        else:
            break

def PDF_SearchAndDecrypt():

    def DecryptPDF(encryptedPath):
        protectedPDF = pikepdf.open(encryptedPath, PASSWORD, allow_overwriting_input = True)
        protectedPDF.save(encryptedPath.replace('ENCRYPTED', 'DECRYPTED'))
        protectedPDF.close()

    fileNames = listdir(PDF_STORAGE_DIR)
    currentDateString = CURRENT_DATE.strftime("%m-%d-%Y")
    issuanceDataPresent = False
    dailyBalanceDataPresent = False
    for fileName in fileNames:
        if '[ENCRYPTED]' and currentDateString in fileName:
            if 'Issuance' in fileName:
                issuanceDataPresent = True
            elif 'DailyBalance'in fileName:
                dailyBalanceDataPresent = True
            DecryptPDF(PDF_STORAGE_DIR + fileName)
            remove(PDF_STORAGE_DIR + fileName)

    if issuanceDataPresent or dailyBalanceDataPresent:
        return issuanceDataPresent, dailyBalanceDataPresent
    else:
        print(f"No Files Found For {CURRENT_DATE}.")
        sleep(2)
        print('Program Exiting Safely')
        sleep(2)
        exit()

def ExtractPDFData(fileName):

    with open(PDF_STORAGE_DIR + fileName, 'rb') as decryptedPDF:
                reader = PdfFileReader(decryptedPDF)
                catalog = reader.trailer["/Root"]
                fileNames = catalog['/Names']['/EmbeddedFiles']['/Names']
                attachments = {}
                for f in fileNames:
                    if isinstance(f, str):
                        name = f
                        dataIndex = fileNames.index(f) + 1
                        fDict = fileNames[dataIndex].getObject()
                        fData = fDict['/EF']['/F'].getData()
                        attachments[name] = fData

    return attachments['part(0)'].decode("utf-8")

def GetRawData(reportType):

    fileNames = listdir(PDF_STORAGE_DIR)
    currentDateString = CURRENT_DATE.strftime("%m-%d-%Y")

    if reportType == 'Issuance':
        fileTag = '[DECRYPTED] Issuance'
    elif reportType == 'DailyBalance':
        fileTag = '[DECRYPTED] DailyBalance'

    for fileName in fileNames:
        if (fileTag in fileName) and (currentDateString in fileName):
            print(f'Extracting Data From: {fileName}')
            return ExtractPDFData(fileName)

def IssuanceData_ParserAndCleaner(issuanceData):

    data = []
    numLines = 0
    for line in issuanceData.split('\n'):
        if len(line) > 0:
            data.append(line.replace('\r', '').replace(",,,,", "").split(","))
            numLines += 1

    if numLines > 1:
        completeData = DataFrame(data = data[1:], columns = data[0][:-4])
        completeData.drop(['$$$$$$$$$$', '$$$$$$$$$$', '$$$$$$$$$$'], axis = 1, inplace = True)
        dealerNames = {'$$$$$$$$$$' : '$$$$$$$$$$', '$$$$$$$$$$' : '$$$$$$$$$$', '$$$$$$$$$$' : '$$$$$$$$$$', '$$$$$$$$$$' : '$$$$$$$$$$'}
        brokers = [dealerNames[broker] for broker in completeData['Dealer']]
        completeData.insert(loc = 7, column = 'Broker', value = brokers)
        return completeData
    else:
        print(f'No Commercial Paper Was Issued For {CURRENT_DATE}.')
        sleep(2)
        return None

def DailyBalanceData_ParserAndCleaner(dailyBalanceData):

    data = []
    for line in dailyBalanceData.split('\n'):
        if len(line) > 0:
            data.append(line.replace('\r', '').replace(",,,,", "").split(","))

    return data[1:]

def UpdateCPDatabase(completeData):

    sheetNames = CPDATABASE_WORKBOOK.sheetnames
    CompleteCPReport_Sheet = CPDATABASE_WORKBOOK[sheetNames[1]]
    currentRowIndex = int(CompleteCPReport_Sheet['P1'].value)

    completeArray = completeData.values
    nrow, ncol = completeArray.shape

    for j in range(ncol):
        cellFormat = 'General'
        isNumber = False
        if j in {2, 8}:
            cellFormat = 'mm-dd-yy'
        elif j in {9, 10, 11}:
            cellFormat = r'$ #,###.00;[red]$ (#,###.00);$ 0.00;'
            isNumber = True
        elif j == 12:
            cellFormat = '0.00'
            isNumber = True

        for i in range(nrow):
            cell = CompleteCPReport_Sheet.cell(row = currentRowIndex + i, column = j + 1)
            if isNumber:
                cell.value = float(completeArray[i, j])
            else:
                cell.value = completeArray[i, j]
            cell.alignment = Alignment(horizontal = 'center')
            cell.number_format = cellFormat

    currentRowIndex += nrow
    CompleteCPReport_Sheet['P1'].value = currentRowIndex
    CompleteCPReport_Sheet.freeze_panes = CompleteCPReport_Sheet['A2']
    if currentRowIndex > 30:
        CompleteCPReport_Sheet.row_dimensions.group(2, currentRowIndex - 25, hidden = True)

    MfCPReport_Sheet = CPDATABASE_WORKBOOK[sheetNames[0]]
    currentRowIndex = int(MfCPReport_Sheet['M1'].value)

    mfColnames = '$$$$$$$$$$'
    mfArray = completeData[mfColnames].values
    nrow, ncol = mfArray.shape

    for j in range(ncol):
        cellFormat = 'General'
        cellAlignment = 'left' if j in {3, 4, 5} else 'right'
        isNumber = False
        isDate = False
        if j in {0, 7, 8}:
            cellFormat = r'$ #,###.00;[red]$ (#,###.00);$ 0.00;'
            isNumber = True
        elif j in {1, 2}:
            cellFormat = 'mm-dd-yy'
            isDate = True
        elif j == 6:
            cellFormat = '0.00'
            isNumber = True

        for i in range(nrow):
            cell = MfCPReport_Sheet.cell(row = currentRowIndex + i, column = j + 1)
            if isNumber:
                cell.value = float(mfArray[i, j])
            elif isDate:
                cell.value = datetime.strptime(mfArray[i, j], '%m/%d/%Y').date()
            else:
                cell.value = mfArray[i, j]
            cell.alignment = Alignment(horizontal = cellAlignment)
            cell.number_format = cellFormat

    currentRowIndex += nrow
    MfCPReport_Sheet['M1'].value = currentRowIndex
    MfCPReport_Sheet.freeze_panes = MfCPReport_Sheet['A2']
    if currentRowIndex > 30:
        MfCPReport_Sheet.row_dimensions.group(2, currentRowIndex - 25, hidden = True)

    DateLogSheet = CPDATABASE_WORKBOOK['Date_Log']
    currentRowIndex = int(DateLogSheet['D1'].value)
    DateLogSheet.cell(row = currentRowIndex, column = 1).value = CURRENT_DATE
    DateLogSheet['D1'].value = currentRowIndex + 1
    DateLogSheet.freeze_panes = DateLogSheet['A2']
    if currentRowIndex > 30:
        DateLogSheet.row_dimensions.group(2, currentRowIndex - 25, hidden = True)

    CPDATABASE_WORKBOOK.save(ISSUANCE_DB_PATH)
    print('CP Database Updated For ' + CURRENT_DATE.strftime("%Y_%m_%d") + ' Successfully.')

    xl = Dispatch("Excel.Application")
    xl.Visible = True
    wb = xl.Workbooks.Open(ISSUANCE_DB_PATH)

def DepositDailyBalanceReport(reportData):

    DATABASE_WORKBOOK = load_workbook(DAILY_BALANCES_DB_PATH)
    sheetNames = DATABASE_WORKBOOK.sheetnames
    cwSheet = DATABASE_WORKBOOK[sheetNames[0]]
    pwSheet = DATABASE_WORKBOOK[sheetNames[1]]

    nrow = len(reportData) ; ncol = len(reportData[0])
    for j in range(ncol):
        cellFormat = 'General'
        isNumber = False
        if j < 4:
            pass
        elif j == 4:
            cellFormat = 'mm-dd-yy'
        elif j < 16:
            cellFormat = r'$ #,###.00;[red]$ (#,###.00);$ 0.00;'
            isNumber = True
        else:
            cellFormat = '0.00'
            isNumber = True

        for i in range(nrow):
            cwCell = cwSheet.cell(row = i + 2, column = j + 1)
            pwCell = pwSheet.cell(row = i + 2, column = j + 1)
            if isNumber:
                pwCell.value = float(cwCell.value)
                cwCell.value = float(reportData[i][j])
            else:
                pwCell.value = cwCell.value
                cwCell.value = reportData[i][j]
            pwCell.alignment = Alignment(horizontal = 'center')
            cwCell.alignment = Alignment(horizontal = 'center')
            pwCell.number_format = cellFormat
            cwCell.number_format = cellFormat

    maxCWCell = cwSheet.cell(row = 1, column = 21)
    maxCWRow = int(maxCWCell.value)
    maxPWCell = pwSheet.cell(row = 1, column = 21)
    maxPWRow = int(maxPWCell.value)

    if nrow + 1 >= maxCWRow:
        pass
    else:
        cwSheet.delete_rows(idx = nrow + 2, amount = maxCWRow - nrow - 1)

    if maxCWRow >= maxPWRow:
        pass
    else:
        pwSheet.delete_rows(idx = maxCWRow + 1, amount = maxPWRow - maxCWRow)

    maxCWCell.value = nrow + 1
    maxPWCell.value = maxCWRow

    DATABASE_WORKBOOK.save(DAILY_BALANCES_DB_PATH)
    print('Daily_Balance_' + CURRENT_DATE.strftime("%Y_%m_%d") + '.xlsx' + ' Deposited Successfully.')

def Clean_CitiPDF_Directory():

    if CURRENT_DATE.weekday() == 3:

        citiPDFnames = listdir(PDF_STORAGE_DIR)
        cutoff = 60*60*24*22

        for pdfName in citiPDFnames:
            fullPath = PDF_STORAGE_DIR + '/' + pdfName
            modifyDate = datetime.fromtimestamp(getmtime(fullPath)).date()
            timeDelta = (CURRENT_DATE - modifyDate).total_seconds()
            if timeDelta > cutoff:
                remove(fullPath)

#Constant Location Variables
PDF_STORAGE_DIR = '$$$$$$$$$$'
DEPENDENCIES_DIR = '$$$$$$$$$$'
ISSUANCE_DB_PATH = '$$$$$$$$$$'
DAILY_BALANCES_DB_PATH = '$$$$$$$$$$'

#Program Execution
try:
    from subprocess import check_call, DEVNULL
    from pkg_resources import working_set
    from sys import executable, exit
    AcquireDependencies()
    from os import getlogin, listdir, rename, remove, getcwd
    from os.path import isfile, getmtime
    from time import sleep
    from datetime import datetime
    from traceback import print_exc
    import pikepdf
    from PyPDF2 import PdfFileReader, PdfFileWriter
    from pandas import DataFrame
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.styles.colors import Color
    from win32com.client import Dispatch
    PASSWORD = AcquirePDFPassword()
    InitializeInbox()
    CURRENT_DATE = AcquireQueryDate()
    CPDATABASE_WORKBOOK = CheckDateLog()
    AcquireIssuanceReport()
    issuanceDataPresent, dailyBalanceDataPresent = PDF_SearchAndDecrypt()

    if issuanceDataPresent:
        issuanceData = GetRawData(reportType = 'Issuance')
        completeData = IssuanceData_ParserAndCleaner(issuanceData)
        if not isinstance(completeData, type(None)):
            UpdateCPDatabase(completeData)

    if dailyBalanceDataPresent:
        dailyBalanceData = GetRawData(reportType = 'DailyBalance')
        reportData = DailyBalanceData_ParserAndCleaner(dailyBalanceData)
        DepositDailyBalanceReport(reportData)

    Clean_CitiPDF_Directory()

except SystemExit:
    print('Program Terminated.')
except Exception:
    print_exc()
